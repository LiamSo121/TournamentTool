import os
import openpyxl
import csv
import itertools
from Color_Helper import Color_Helper



class Shlush():
    def __init__(self) -> None:
        self.color_helper = Color_Helper()       


    def shlush_to_csv(self,excel_name):
        # CSV
        wb = openpyxl.load_workbook("separated_values.xlsx")

        # get the name of the first sheet
        sheet_name = wb.sheetnames[0]

        # get the first sheet
        sheet = wb[sheet_name]

        # create a new CSV file
        csv_file_name = f"{excel_name}_Brands_ToImport.csv"
        csv_file = open(csv_file_name, "w")

        exists = {}
        row_number = 1
        # loop through the rows in the new Excel file
        for row in sheet:
            # get the values from Each row
            partner_id = row[0].value
            certification = row[1].value
            exclude_event = row[2].value
            exclude_welcome = row[2].value
            allow_opt_out = row[3].value

            # Fix for int to str on empty spaces
            if partner_id in ['nan', 'None'] or partner_id is None:
                continue

            # Validation of Certifications in different names and changing them to script convention.
            if certification == 'Curacao-Cert':
                certification = 'MGA'
            elif certification == 'Curaçao':
                certification = 'Curacao'
            elif certification == 'EST. MGA':
                certification = 'Curacao'

            # Check Cert
            if certification == 'Curacao':
                certification = 'FALSE'
            elif certification == 'MGA':
                certification = 'TRUE'
            elif certification == 'CU':
                certification = 'FALSE'
            elif certification == 'Curacao':
                certification = 'FALSE'
            elif certification is None or certification == 'None':
                certification = 'FALSE'

            # Error Handling certification
            if certification not in ('TRUE', 'FALSE'):
                self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, Certification = '{certification}'")

            # Check exclude_event
            if exclude_event == 'Opt out Rush Hour':
                exclude_event = 'TRUE'
            elif exclude_event == 'Opt out ALL':
                exclude_event = 'TRUE'
            elif exclude_event == 'Opt in ALL':
                exclude_event = 'FALSE'

            # Error Handling exclude_event
            if exclude_event not in ('TRUE', 'FALSE'):
                self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, ExcludeEvent = '{exclude_event}'")

            # Check Exclude Welcome
            if exclude_welcome == 'Opt out Welcome Bonus':
                exclude_welcome = 'TRUE'
            elif exclude_welcome == 'Opt out ALL':
                exclude_welcome = 'TRUE'
            elif exclude_welcome == 'Opt in ALL':
                exclude_welcome = 'FALSE'

            # Error Handling exclude_welcome
            if exclude_welcome not in ('TRUE', 'FALSE'):
                self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, ExcludeWelcome = '{exclude_welcome}'")

            if allow_opt_out == 'Yes':
                allow_opt_out = 'TRUE'
            elif allow_opt_out == 'No':
                allow_opt_out = 'FALSE'
            elif allow_opt_out == 'Pending':
                allow_opt_out = 'FALSE'

            # Error Handling allow_opt_out
            if allow_opt_out not in ('TRUE', 'FALSE'):
                self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, AllowOptOut = '{allow_opt_out}'")

            # Write final line to CSV row
            if partner_id not in exists:
                csv_file.write(f"{partner_id};{certification};{exclude_event};{exclude_welcome};;{allow_opt_out}\n")
                row_number += 1
                if partner_id in ["eduardo", "betplatino-new", "30758", "stage2", "31076", "31232", "31439", "3106", "1462",
                                "30480",
                                "30050", "1453", "1740", "30032", "30086", "72", "30181", "mildcasino", "101549",
                                "888bits",
                                "30631", "1000138", "3004612", "3004542", "30103"]:
                    # Error Handling certification
                    if certification not in ('TRUE', 'FALSE'):
                        self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, Certification = '{certification}'")
                    # Error Handling exclude_event
                    if exclude_event not in ('TRUE', 'FALSE'):
                        self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, ExcludeEvent = '{exclude_event}'")
                    # Error Handling exclude_welcome
                    if exclude_welcome not in ('TRUE', 'FALSE'):
                        self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, ExcludeWelcome = '{exclude_welcome}'")
                    # Error Handling allow_opt_out
                    if allow_opt_out not in ('TRUE', 'FALSE'):
                        self.color_helper.red_colored_printer(f"Issue in CSV line: {row_number}, AllowOptOut = '{allow_opt_out}'")
                    csv_file.write(f"n-{partner_id};{certification};{exclude_event};{exclude_welcome};;{allow_opt_out}\n")
                    row_number += 1
                exists[partner_id] = True

        # close the CSV file
        csv_file.close()
        # Remove Separated Values
        os.remove("separated_values.xlsx")



    def shlush_brands_func(self,excel_name):
        path_from_input = f"{excel_name}_Brands_Output.xlsx"
        wb = openpyxl.load_workbook(path_from_input)
        ws = wb.active
        # Variables to store column numbers
        partner_id_col = brand_certification_col = welcome_bonus_rush_hour_col = allow_opt_out_col = None
        # Extract column numbers
        for cell in ws[3]:  # 3rd row
            if cell.value == 'PartnerId':
                partner_id_col = cell.column
            elif cell.value == 'Brand Certification':
                brand_certification_col = cell.column
            elif cell.value == 'Welcome Bonus + Rush Hour':
                welcome_bonus_rush_hour_col = cell.column
            elif cell.value == 'AllowOptOut':
                allow_opt_out_col = cell.column
        # Extract and explode values from specified columns, skip first 3 rows
        exploded_values = []
        for row_num in range(4, ws.max_row + 1):  # start from 4th row, 1-based index
            row_values = [
                str(ws.cell(row=row_num, column=partner_id_col).value or '').split(','),
                str(ws.cell(row=row_num, column=brand_certification_col).value or '').split(','),
                str(ws.cell(row=row_num, column=welcome_bonus_rush_hour_col).value or '').split(','),
                str(ws.cell(row=row_num, column=allow_opt_out_col).value or '').split(','),
            ]
            # Ignore the entire row if one of the values is None
            if any('' in values for values in row_values):
                continue
            row_combinations = list(itertools.product(*row_values))
            exploded_values.extend(row_combinations)

        # Write exploded values to new Excel file, removing duplicates
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        written_rows = set()

        for exploded_row in exploded_values:
            row_to_write = [x.strip() for x in exploded_row]
            row_tuple = tuple(row_to_write)
            if row_tuple not in written_rows:
                ws_new.append(row_to_write)
                written_rows.add(row_tuple)

        wb_new.save('separated_values.xlsx')

    def shlush_games_func(self,excel_name):
        path_from_input = f"{excel_name}_Games_Output.xlsx"
        wb = openpyxl.load_workbook(path_from_input)
        ws = wb.active

        # Extract column number where name is 'Game Code'
        col_num = None
        for cell in ws[3]:  # 3rd row since we skip the first two rows
            if cell.value in ['Game Code', 'BrandedGame']:
                col_num = cell.column
                break

        # Extract values from col_num, skip first 3 rows
        values = []
        for row_num in range(4, ws.max_row + 1):  # start from 4th row, 1-based index
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                values.append(cell_value)

        # Remove duplicates
        values = list(set(values))
        csv_file_name = f"{excel_name}_ToImport.csv"

        # Write values to CSV file
        with open(csv_file_name, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            csv_line = 1
            for value in values:
                # split value by ',' and write each part to a new row
                for part in value.split(','):
                    part_with_semicolons = part.strip() + ';;;;'
                    if not part_with_semicolons.startswith('SlotMachine_') or '_' not in part_with_semicolons:
                        self.color_helper.red_colored_printer(f'Issue in CSV line:{csv_line}, GameCode not according to format: {part}')
                    csv_line += 1
                    writer.writerow([part_with_semicolons])
